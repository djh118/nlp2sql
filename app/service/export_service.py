import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _char_width(s: str) -> int:
    w = 0
    for ch in s:
        w += 2 if re.match(r"[\u4e00-\u9fff]", ch) else 1
    return w


def _to_cell_value(v):
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        v = v.strip()
        if not v:
            return v
        try:
            return int(v)
        except ValueError:
            pass
        try:
            return float(v)
        except ValueError:
            pass
    return v


def rows_to_excel(rows: list[dict], sheet_name: str = "数据导出") -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([_to_cell_value(row.get(h, "")) for h in headers])

        for ci, h in enumerate(headers, 1):
            col_letter = get_column_letter(ci)
            max_w = _char_width(str(h))
            for row in ws.iter_rows(min_col=ci, max_col=ci, values_only=True):
                cell_w = _char_width(str(row[0] or ""))
                if cell_w > max_w:
                    max_w = cell_w
            ws.column_dimensions[col_letter].width = max_w + 4
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
